# -*- coding: utf-8 -*-
import os
import requests
import json
from bs4 import BeautifulSoup
from dataclasses import dataclass
import re
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from PySide6.QtWidgets import (QApplication, QDialog, QListWidget, QListWidgetItem, QPushButton, 
                               QVBoxLayout, QHBoxLayout, QLabel, QFileDialog, QLineEdit, QSpinBox, QMessageBox)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
}

@dataclass
class All_url:
    name: str
    time: str
    url: str

@dataclass
class Post_data:
    name: str
    data: str


def read_excel_rows_by_headers(header_names: list[str], usersDB_path: str) -> list[list]:
    """从 Excel 表格第一行匹配列名，将匹配列按行对应写入多维列表。"""
    if not os.path.isfile(usersDB_path):
        raise FileNotFoundError(f'用户数据库文件不存在: {usersDB_path}')

    workbook = load_workbook(usersDB_path, read_only=True, data_only=True)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    normalized_headers = [str(cell).strip().lower() if cell is not None else '' for cell in first_row]

    header_indices = []
    for header_name in header_names:
        if header_name is None:
            continue
        target = str(header_name).strip().lower()
        if target in normalized_headers:
            header_indices.append(normalized_headers.index(target))

    if not header_indices:
        return []

    result = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        result.append([row[index] for index in header_indices])

    return result


def get_cookies(url):
    response = requests.get(url, headers=headers)
    cookies = response.cookies.get_dict()
    print(cookies)
    return cookies


def get_data_from_url(post_url, base_url, url, cookies, data):

    # 构建请求头，包含cookie和其他必要的字段
    headers_get = {
    'Host': f'{base_url}',
    'Cookie': f'acw_tc={cookies["acw_tc"]}; aliyungf_tc={cookies["aliyungf_tc"]}; PHPSESSID={cookies["PHPSESSID"]}',
    'Sec-Ch-Ua': '"-Not.A/Brand";v="8", "Chromium";v="102"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Upgrade-Insecure-Requests': '1',
    'Dnt': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Dest': 'iframe',
    'Referer': f'{url}',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'close'
    }
    headers_post = {
    'Host': f'{base_url}',
    'Cookie': f'acw_tc={cookies["acw_tc"]}; aliyungf_tc={cookies["aliyungf_tc"]}; PHPSESSID={cookies["PHPSESSID"]}',
    'Sec-Ch-Ua': '"-Not.A/Brand";v="8", "Chromium";v="102"',
    'Dnt': '1',
    'Sec-Ch-Ua-Mobile': '?0',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Accept': '*/*',
    'X-Requested-With': 'XMLHttpRequest',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Origin': f'https://{base_url}/',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Dest': 'empty',
    'Referer': f'{url}?from_device=mobile',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'close'
    }
    post_response = requests.post(post_url, headers=headers_post , data=data)
    # print(post_response.text)
    if post_response.json().get('errNo') == 100:
        print('\033[91m' + f"\n{data}数据无查询结果，已跳过" + '\033[0m')
        return "NotFound"
    response = requests.get(f"https://v3uehkhd.yichafen.com/public/queryresult/from_device/mobile.html", headers=headers_get)
    # print(response.text)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'lxml')
    table = soup.find('table', class_='table table-bordered s_table-bordered js_result_table')
    result = []
    if table:
        rows = table.find_all('tr')
        for row in rows:
            cols = row.find_all('td')
            cols = [col.text.strip() for col in cols]
            result.append(cols)
    # print(result)
    return result

def get_url_list_from_menu(menu_url :str):
    print(f"\033[1;32;40m正在获取查询界面列表...\033[0m")
    response = requests.get(menu_url,headers=headers)
    response.encoding = 'utf-8'
    # print(response.text)
    soup = BeautifulSoup(response.text, 'lxml')
    all_url_list = []
    for a_type in soup.find_all('a', class_='weui-cell weui-cell_access'):
        url=a_type.get('href')
        p_data = a_type.find_all('p')
        name = p_data[0].text.strip()
        time = p_data[1].text
        all_url_list.append(All_url(name, time, url))
    print(all_url_list)
    return all_url_list

def get_post_data(url :str ,usersDB_path :str = "./data.xlsx"):
    '''从查询界面url获取POST_URL,POST_DATA'''
    # 获取post_url
    response = requests.get(url,headers=headers)
    response.encoding = 'utf-8'
    # print(response.text)
    soup = BeautifulSoup(response.text, 'lxml')
    script_content = soup.find('script',type='text/javascript',src=False).string
    post_url :str = re.search(r'\$\.post\(\s*(["\'])([^"\']+)\1', script_content)
    if post_url:
        post_url = post_url.group(2)
    else:
        print("\033[91m未找到post_url\033[0m")
        raise RuntimeError("未找到post_url")
    print("\033[1;32;40m" + f"获取到post_url:{post_url}" + "\033[0m")

    # 获取post_data
    post_data = []
    input_content = soup.find_all('input')
    print(input_content)
    for input_tag in input_content:
        name=input_tag.get("name")
        data_name=input_tag.get("data-sname")
        post_data.append(Post_data(name, data_name))
        #从文件中获取数据

    if not post_data:
        raise RuntimeError("未找到post_data")
    print("\033[1;32;40m" + f"获取到post_data:{post_data}" + "\033[0m")
    return post_url, post_data
    
def save_data_to_excel(data, save_path):
    """将数据保存到 Excel"""
    if not data:
        return

    headers = [str(item[0]).strip() for item in data if len(item) >= 2]
    values = [item[1] for item in data if len(item) >= 2]

    if not headers:
        return

    save_dir = os.path.dirname(save_path)
    if save_dir and not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)

    if os.path.exists(save_path):
        workbook = load_workbook(save_path)
        sheet = workbook.active

        existing_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))]
        if not any(existing_headers):
            existing_headers = []
            sheet.delete_rows(1, sheet.max_row)

        header_positions = {}
        for idx, header in enumerate(existing_headers):
            if header:
                header_positions.setdefault(header, []).append(idx)

        for header in headers:
            if header not in header_positions:
                existing_headers.append(header)
                header_positions.setdefault(header, []).append(len(existing_headers) - 1)
                sheet.cell(row=1, column=len(existing_headers), value=header)

        row_data = [None] * len(existing_headers)
        for header, value in zip(headers, values):
            positions = header_positions.get(header, [])
            target_idx = None
            for idx in positions:
                if idx >= len(row_data) or row_data[idx] in (None, ''):
                    target_idx = idx
                    break

            if target_idx is None:
                existing_headers.append(header)
                target_idx = len(existing_headers) - 1
                header_positions.setdefault(header, []).append(target_idx)
                sheet.cell(row=1, column=len(existing_headers), value=header)
                row_data.append(None)

            if target_idx >= len(row_data):
                row_data.extend([None] * (target_idx + 1 - len(row_data)))
            row_data[target_idx] = value

        sheet.append(row_data)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        sheet.append(values)

    workbook.save(save_path)


def select_query_url(all_url_list, config):
    """显示 PySide6 窗口选择查询对象，返回选中的索引和 url"""
    app = QApplication.instance() or QApplication(sys.argv)
    
    dialog = QDialog(None)
    dialog.setWindowTitle('查询界面选择')
    dialog.setGeometry(100, 100, 600, 450)
    dialog.setStyleSheet("""
        QDialog {
            background-color: #f0f0f0;
        }
        QLabel {
            color: #333;
            font-size: 12pt;
            font-weight: bold;
        }
        QListWidget {
            border: 1px solid #ddd;
            border-radius: 4px;
            background-color: white;
            font-size: 11pt;
        }
        QPushButton {
            background-color: #0078d4;
            color: white;
            border: none;
            border-radius: 4px;
            padding: 6px 12px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: #1084d9;
        }
        QPushButton:pressed {
            background-color: #106ebe;
        }
    """)
    
    layout = QVBoxLayout()
    
    # 添加标签
    label = QLabel('请选择爬取对象：')
    layout.addWidget(label)
    
    # 添加列表框
    list_widget = QListWidget()
    for i, url in enumerate(all_url_list):
        item_text = f"{i+1}. {url.name} (日期: {url.time})"
        list_widget.addItem(item_text)
    
    list_widget.setCurrentRow(0)
    layout.addWidget(list_widget)
    
    # 添加按钮
    button_layout = QHBoxLayout()
    ok_button = QPushButton('确定')
    settings_button = QPushButton('⚙️设置')
    cancel_button = QPushButton('取消')
    button_layout.addWidget(ok_button)
    button_layout.addWidget(settings_button)
    button_layout.addWidget(cancel_button)
    layout.addLayout(button_layout)
    
    dialog.setLayout(layout)
    
    # 结果变量
    result = {'idx': None, 'url': None, 'settings': False}
    
    def on_ok():
        current_row = list_widget.currentRow()
        if current_row >= 0:
            result['idx'] = current_row
            result['url'] = all_url_list[current_row].url
            dialog.accept()
    
    def on_settings():
        if settings_dialog(config):
            result['settings'] = True
            dialog.accept()
    
    def on_cancel():
        dialog.reject()
    
    ok_button.clicked.connect(on_ok)
    settings_button.clicked.connect(on_settings)
    cancel_button.clicked.connect(on_cancel)
    
    if dialog.exec() == QDialog.Accepted:
        if result['settings']:
            return 'settings', None
        return result['idx'], result['url']
    else:
        return None, None


def select_save_path(default_name):
    """显示 PySide6 文件保存对话框，返回选中的路径"""
    app = QApplication.instance() or QApplication(sys.argv)
    
    file_dialog = QFileDialog(None)
    file_dialog.setWindowTitle('选择保存路径')
    file_dialog.setDefaultSuffix('xlsx')
    file_dialog.setNameFilters(['Excel 文件 (*.xlsx)', '所有文件 (*)'])
    file_dialog.setFileMode(QFileDialog.AnyFile)
    file_dialog.setAcceptMode(QFileDialog.AcceptSave)
    file_dialog.selectFile(default_name)
    
    if file_dialog.exec() == QFileDialog.Accepted:
        files = file_dialog.selectedFiles()
        return files[0] if files else None
    else:
        return None


def settings_dialog(config):
    """显示设置窗口，允许用户配置 base_url、usersDB_path 和 num_threads"""
    app = QApplication.instance() or QApplication(sys.argv)
    dialog = QDialog(None)
    dialog.setWindowTitle('程序设置')
    dialog.setGeometry(100, 100, 500, 350)
    dialog.setStyleSheet("""
        QDialog {
            background-color: #f0f0f0;
        }
        QLabel {
            color: #333;
            font-size: 11pt;
        }
        QLineEdit {
            border: 1px solid #ddd;
            border-radius: 4px;
            padding: 6px;
            background-color: white;
            font-size: 11pt;
        }
        QSpinBox {
            border: 1px solid #ddd;
            border-radius: 4px;
            padding: 6px;
            background-color: white;
            font-size: 11pt;
        }
        QPushButton {
            background-color: #0078d4;
            color: white;
            border: none;
            border-radius: 4px;
            padding: 6px 12px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: #1084d9;
        }
        QPushButton:pressed {
            background-color: #106ebe;
        }
    """)
    
    layout = QVBoxLayout()
    
    # Base URL 设置
    base_url_label = QLabel('Base URL:')
    base_url_input = QLineEdit()
    base_url_input.setPlaceholderText('从yichafen主页网址获取,例如：xxx.yichafen.com')
    base_url_input.setText(config.get('base_url', ''))
    layout.addWidget(base_url_label)
    layout.addWidget(base_url_input)
    
    # 用户数据库路径设置
    db_path_label = QLabel('用户数据库路径:')
    db_path_input = QLineEdit()
    db_path_input.setPlaceholderText('请选择本地 Excel 文件路径')
    db_path_input.setText(config.get('usersDB_path(excel)', ''))
    db_path_layout = QHBoxLayout()
    db_path_layout.addWidget(db_path_input)
    db_path_browse_button = QPushButton('选择文件')
    db_path_layout.addWidget(db_path_browse_button)
    layout.addWidget(db_path_label)
    layout.addLayout(db_path_layout)
    
    # 线程数设置
    threads_label = QLabel('爬取线程数:')
    threads_spinbox = QSpinBox()
    threads_spinbox.setMinimum(1)
    threads_spinbox.setMaximum(16)
    threads_spinbox.setValue(config.get('num_threads', 4))
    layout.addWidget(threads_label)
    layout.addWidget(threads_spinbox)
    
    # 添加按钮
    button_layout = QHBoxLayout()
    ok_button = QPushButton('保存')
    cancel_button = QPushButton('取消')
    button_layout.addWidget(ok_button)
    button_layout.addWidget(cancel_button)
    layout.addLayout(button_layout)
    
    dialog.setLayout(layout)
    
    # 结果变量
    result = {'saved': False}
    
    def on_browse_db_path():
        file_path, _ = QFileDialog.getOpenFileName(dialog, '选择用户数据库文件', '', 'Excel 文件 (*.xlsx);;所有文件 (*)')
        if file_path:
            db_path_input.setText(file_path)
    
    db_path_browse_button.clicked.connect(on_browse_db_path)
    
    def on_ok():
        base_url = base_url_input.text().strip()
        db_path = db_path_input.text().strip()
        num_threads = threads_spinbox.value()
        
        if not base_url or not db_path:
            QMessageBox.warning(dialog, '提示', '请填写所有必要的配置信息')
            return
        
        # 更新 config
        config['base_url'] = base_url
        config['usersDB_path(excel)'] = db_path
        config['num_threads'] = num_threads
        
        # 保存到文件
        try:
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
            result['saved'] = True
            dialog.accept()
        except Exception as e:
            QMessageBox.critical(dialog, '错误', f'保存配置失败：{e}')
    
    def on_cancel():
        dialog.reject()
    
    ok_button.clicked.connect(on_ok)
    cancel_button.clicked.connect(on_cancel)
    
    if dialog.exec() == QDialog.Accepted and result['saved']:
        return True
    else:
        return False


def main():
    
    app = QApplication.instance() or QApplication(sys.argv)
    #读取配置文件
    try:
        with open('config.json', 'r', encoding='utf-8') as f:  
            config = json.load(f)
    except FileNotFoundError:
        print("\033[91m未找到配置文件config.json，创建默认配置\033[0m")
        config = {
            'base_url': '',
            'usersDB_path(excel)': '',
            'num_threads': 4
        }
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
    
    # 检查必要的配置，如果缺少则打开设置窗口
    while not config.get('base_url') or not config.get('usersDB_path(excel)'):
        print("需要配置必要的参数，打开设置窗口...")
        if not settings_dialog(config):
            print("用户取消了设置")
            return
        # 重新加载配置
        try:
            with open('config.json', 'r', encoding='utf-8') as f:  
                config = json.load(f)
        except:
            pass
    
    # 从 config 中读取线程数，默认为 4
    num_threads = config.get('num_threads', 4)
    if num_threads < 1:
        num_threads = 4
    # 从网站首页获取查询界面url列表
    base_url = config['base_url']
    menu_url = f'https://{base_url}/'
    all_url_list = get_url_list_from_menu(menu_url)
    if not all_url_list:
        print("\033[91m无法获取查询界面列表\033[0m")
        return

    # 使用 UI 窗口选择查询界面
    while True:
        selected_idx, selected_url = select_query_url(all_url_list, config)
        if selected_idx is None:
            print("用户取消了选择")
            return
        if selected_idx == 'settings':
            try:
                with open('config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except Exception:
                pass
            if not config.get('base_url') or not config.get('usersDB_path(excel)'):
                print("用户取消了设置")
                return
            base_url = config['base_url']
            menu_url = f'https://{base_url}/'
            all_url_list = get_url_list_from_menu(menu_url)
            if not all_url_list:
                print("\033[91m无法获取查询界面列表\033[0m")
                return
            continue
        break
    
    url = f"https://{base_url}{selected_url}"
    
    # 使用文件保存对话框选择保存路径
    default_filename = all_url_list[selected_idx].name + ".xlsx"
    save_path = select_save_path(default_filename)
    if not save_path:
        print("用户取消了保存路径选择")
        return
    
    # 从查询界面url获取POST_URL,POST_DATA需求and cookies
    try:
        post_url, post_fields = get_post_data(url,config['usersDB_path(excel)'])
        cookies = get_cookies(url)
    except RuntimeError as e:
        print('\033[91m'+f"运行时出现错误：{e}"+'\033[0m')
        return
    except Exception as e:
        print('\033[91m'+f"错误：{e}"+'\033[0m')
        return
    post_url = f"https://{base_url}{post_url}"

    # 从本地根据 POST_DATA 需求获取数据
    data_fields = [field for field in post_fields if field.data]
    field_names = [field.data for field in data_fields]
    try:
        excel_rows = read_excel_rows_by_headers(field_names, config['usersDB_path(excel)'])
    except Exception as e:
        print('\033[91m'+f"读取本地用户数据库失败：{e}"+'\033[0m')
        return
    if excel_rows:
        print('\033[1;32;40m' + '找到匹配的本地用户数据：' + '\033[0m')
        print(excel_rows)
    else:
        print('\033[93m' + '未从本地用户数据库读取到匹配数据，或未找到对应列名。' + '\033[0m')
        return


    # 开始爬取数据
    success_count = 0
    failed_count = 0
    error_messages = []
    save_lock = Lock()
    
    # 预先生成与线程数相同的 cookies
    print(f"正在为 {num_threads} 个线程预生成 cookies...")
    thread_cookies_list = []
    for i in range(num_threads):
        try:
            cookies = get_cookies(url)
            thread_cookies_list.append(cookies)
        except Exception as e:
            print(f"获取第 {i+1} 个 cookies 失败：{e}")
            # 如果获取失败，使用前一个或继续
            if thread_cookies_list:
                thread_cookies_list.append(thread_cookies_list[-1])
    
    if not thread_cookies_list:
        print("\033[91m无法获取任何 cookies，程序退出\033[0m")
        return
    
    print(f"成功预生成 {len(thread_cookies_list)} 个 cookies\n")
    
    def fetch_and_save_data(row_idx, row, thread_id):
        """线程工作函数：获取数据并保存"""
        nonlocal success_count, failed_count
        
        # 为此线程分配对应的 cookies
        thread_cookies = thread_cookies_list[thread_id % len(thread_cookies_list)]
        
        data = {data_fields[i].name: row[i] for i in range(min(len(row), len(data_fields)))}
        
        max_retries = 3
        return_data = None
        for attempt in range(max_retries):
            try:
                return_data = get_data_from_url(post_url, base_url, url, thread_cookies, data)
                break
            except (ConnectionError, ConnectionResetError) as e:
                if attempt < max_retries - 1:
                    print(f"[线程 {row_idx}] 连接错误（尝试 {attempt + 1}/{max_retries}）：{e}")
                    print(f"[线程 {row_idx}] 正在使用备用 cookies 重试...")
                    # 尝试更换为其他线程的 cookies
                    thread_cookies = thread_cookies_list[(thread_id + 1) % len(thread_cookies_list)]
                    continue
                else:
                    print(f"[线程 {row_idx}] 连接错误，已重试 {max_retries} 次：{e}")
                    failed_count += 1
                    error_messages.append(f"[线程 {row_idx}] 连接错误，已重试 {max_retries} 次：{e}")
                    return
        
        if return_data and return_data != "NotFound":
            try:
                with save_lock:
                    save_data_to_excel(return_data, save_path)
                success_count += 1
            except PermissionError as e:
                print(f'\033[91m[线程 {row_idx}] 保存文件失败：{e}\033[0m')
                print('\033[93m请不要打开表格文件，待数据保存完成后再打开\033[0m')
                failed_count += 1
                error_messages.append(f"[线程 {row_idx}] 保存文件失败：{e}")
            except Exception as e:
                print(f'\033[91m[线程 {row_idx}] 保存文件出错：{e}\033[0m')
                failed_count += 1
                error_messages.append(f"[线程 {row_idx}] 保存文件出错：{e}")
        else:
            if return_data == "NotFound":
                failed_count += 1  # 数据不正确，已跳过
            else:
                failed_count += 1
                error_messages.append(f"[线程 {row_idx}] 未获取到有效返回数据。")
    
    # 使用线程池并发处理
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = {}
        for row_idx, row in enumerate(excel_rows, 1):
            thread_id = (row_idx - 1) % num_threads
            future = executor.submit(fetch_and_save_data, row_idx, row, thread_id)
            futures[future] = row_idx
        
        # 使用 tqdm 显示进度
        for future in tqdm(as_completed(futures), total=len(futures), desc='查询进度', colour='green', position=0, leave=False):
            pass
    
    if success_count == 0 and failed_count == len(excel_rows):
        app = QApplication.instance() or QApplication(sys.argv)
        if error_messages:
            summary = '\n'.join(error_messages[:10])
            if len(error_messages) > 10:
                summary += f"\n...还有 {len(error_messages) - 10} 条错误信息。"
        else:
            summary = '所有请求均未返回有效数据，程序已退出。'
        QMessageBox.critical(None, '错误', f'全部任务均失败，程序已退出。\n\n{summary}')
        return

    print(f'\033[1;32;40m爬取完成：成功 {success_count} 条，失败 {failed_count} 条\033[0m')
    QMessageBox.information(None, '提示', f'爬取完成：成功 {success_count} 条，失败 {failed_count} 条')
    
if __name__ == '__main__':
    main()